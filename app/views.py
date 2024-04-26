# views.py
from django.shortcuts import render, redirect
from django.urls import reverse
from .models import Pechera
from .models import Lavado
from .models import Planta
from .forms import *
import serial
from .forms import FormPechera, FiltroPecheraForm
from django.utils import timezone
from django.http import HttpResponse, HttpResponseRedirect
from datetime import timedelta, datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import JsonResponse
from datetime import date

# Create your views here.


from .models import Pechera

def index(request):
    # Obtén las últimas pecheras (por ejemplo, las últimas 20) ordenadas por fecha de fabricación descendente
    latest_pecheras = Pechera.objects.exclude(eliminada=True).order_by('-fecha_fabricacion')[:10]

    # Calcula la cantidad de pecheras fabricadas el último año
    today = datetime.now()
    last_year = today - timedelta(days=365)
    manufactured_last_year = Pechera.objects.filter(fecha_fabricacion__gte=last_year).count()

    # Calcular la cantidad de pecheras fabricadas el último mes
    last_month = today - timedelta(days=30)
    manufactured_last_month = Pechera.objects.filter(fecha_fabricacion__gte=last_month).count()

    # Pecheras lavadas
    washed_last_month = Lavado.objects.filter(fecha_lavado__gte=last_month).count()

    # Pecheras defectuosas del último mes
    faulty_last_month = Pechera.objects.filter(eliminada=True, fecha_fabricacion__gte=last_month).count()

    # Pecheras disponibles
    disponibles = Pechera.objects.filter(planta__icontains=1,eliminada=False).count()

    # Pecheras en planta
    pecheras_no_eliminadas = Pechera.objects.exclude(eliminada=True)[:]

    # Pecheras en planta
    planta = pecheras_no_eliminadas.exclude(planta__iexact=1).count()

    context = {
        'latest_pecheras': latest_pecheras,
        'manufactured_last_year': manufactured_last_year,
        'manufactured_last_month': manufactured_last_month,
        'faulty_last_month': faulty_last_month,
        'washed_last_month': washed_last_month,
        'disponibles': disponibles,
        'planta': planta,
    }

    return render(request, 'app/index.html', context)

def listado(request):
    # Obtener todas las pecheras ordenadas por fecha de fabricación descendente
    latest_pecheras = Pechera.objects.exclude(eliminada=True).order_by('-fecha_fabricacion')[:]

    # Obtener todas las plantas
    plantas = Planta.objects.all()

    context = {
        'latest_pecheras': latest_pecheras,'plantas': plantas
    }

    return render(request, 'app/listado.html', context)

def plantas(request):
    # Obtener todas las pecheras ordenadas por fecha de fabricación descendente
    latest_pecheras = Pechera.objects.exclude(eliminada=True).order_by('-fecha_fabricacion')[:]

    # Obtener todas las plantas
    plantas = Planta.objects.all()

    # Obtener la cantidad de pecheras por planta y los kilos de plastico por planta
    for planta in plantas:
        planta.cantidad = Pechera.objects.filter(planta__icontains=planta.id_planta,eliminada=False).count()
        planta.kilos = (planta.cantidad*150)/1000

    context = {
        'latest_pecheras': latest_pecheras,
        'plantas': plantas
    }

    return render(request, 'app/plantas.html', context)

def ingreso(request):  
    fecha_actual = date.today()
    id_pechera = " "

    if request.method == 'GET':
        if request.GET.get("id") != None:
            id_pechera = request.GET.get("id")

    # Obtener todas las plantas
    plantas = Planta.objects.all()

    parameters_choices = [
    ('1', 'Nuevo'),
    ('2', 'Perfectas condiciones'),
    ('3', 'Ligermante desgastanda'),
    ('4', 'Desgastada'),]

    opciones_indice = [('Si'),('No'),]

    context = {'fecha_actual':fecha_actual, 'id_pechera':id_pechera,'plantas': plantas, 'parameters_choices': parameters_choices,'opciones_indice': opciones_indice}
    return render(request, 'app/ingreso.html', context)

def ingreso_alerta(request,alerta):   
    fecha_actual = date.today()
    id_pechera = " "
    alerta = alerta

    # Obtener todas las plantas
    plantas = Planta.objects.all()

    parameters_choices = [
    ('1', 'Nuevo'),
    ('2', 'Perfectas condiciones'),
    ('3', 'Ligermante desgastanda'),
    ('4', 'Desgastada'),]

    opciones_indice = [('Si'),('No'),]

    context = {'fecha_actual':fecha_actual,'id_pechera':id_pechera, 'plantas': plantas, 'parameters_choices': parameters_choices,'opciones_indice': opciones_indice,'alerta':alerta}
    return render(request, 'app/ingreso.html', context)

def leer_ingreso(request):
    arduino_port = 'COM4'
    baud_rate = 9600
    arduino = None

    try:
        arduino = serial.Serial(arduino_port, baud_rate)
        while True:
            rfid_data_full = arduino.readline().decode().strip()
            # Verificar si el valor es "Listo para leer etiquetas RFID..."
            if rfid_data_full != "Listo para leer etiquetas RFID...":
                # Extraer el UID eliminando el texto no deseado
                prefix = "Número de serie de la tarjeta RFID:"
                if rfid_data_full.startswith(prefix):
                    rfid_data = rfid_data_full[len(prefix):].strip().replace(" ", "")
                else:
                    rfid_data = rfid_data_full.replace(" ", "")

                if rfid_data:                    
                    print("RFID TAG: ", rfid_data)

                    # Verificar si el UID ya existe en la base de datos
                    existing_pechera = Pechera.objects.filter(id_pechera=rfid_data).first()

                    if existing_pechera:
                        # Notificar al usuario
                        print("UID ya existe en la base de datos")
                        return redirect('ingreso_alerta',alerta="1")
                    else:
                        # Notificar al usuario
                        print("UID no registrado en la base de datos")

                        # Generar la URL de la página 'lectura' con el ID de la pechera como parámetro
                        url = reverse('ingreso') + f'?id={rfid_data}'

                        # Redirigir a la página 'lectura' con el ID de la pechera como parámetro en la URL
                        return redirect(url)            

    except KeyboardInterrupt:
        print("Lectura de RFID detenida.")
    except Exception as e:
        print("Error: ", e)
    finally:
        if arduino is not None:
            arduino.close()

    return HttpResponse("Error en la lectura de datos, retroseda y recargue la pagina. IMPORTANTE: verifique que el dispositivo pueda leer información antes de comenzar la lectura.")


def guardar(request):
    id_pechera = request.POST['id_pechera']

    if id_pechera == " ":  
        return redirect('ingreso_alerta',alerta="2")
    else:
        existing_pechera = Pechera.objects.filter(id_pechera=id_pechera).first()
        if existing_pechera:
            return redirect('ingreso_alerta',alerta="1")
        else:
            Pechera.objects.create(
                id_pechera = id_pechera,
                fecha_fabricacion = request.POST['fecha_fabricacion'],
                talla = request.POST['talla'],
                observaciones = request.POST['observaciones'],
                planta = request.POST['planta'],
                parameters = request.POST['parameters'],
                indice_microbiologico = request.POST['indice_microbiologico']
                )
            print("Pechera creada con exito")
            return redirect('ingreso_alerta',alerta="3")

def lectura(request): 
    pechera = None  # Inicialmente, no se muestra ninguna pechera
    lista_lavados = None  # Inicialmente, no se muestra ninguna pechera
    if request.method == 'GET':
        id_pechera = request.GET.get("id")
        pechera = Pechera.objects.filter(id_pechera=id_pechera).first()
        lista_lavados = Lavado.objects.filter(id_pechera=id_pechera)[:]
        if pechera:
            if pechera.eliminada:
                return redirect('lavado_alerta') 

    context = {'pechera': pechera, "lista_lavados":lista_lavados}

    return render(request, 'app/lectura.html', context)

def lectura_alerta(request):   
    alerta = "UID no registrado en la base de datos"
    pechera = None  # Inicialmente, no se muestra ninguna pechera
    lista_lavados = None  # Inicialmente, no se muestra ninguna pechera
    context = {'pechera': pechera, "lista_lavados":lista_lavados, 'alerta': alerta}
    return render(request, 'app/lectura.html', context)

def leer(request):
    arduino_port = 'COM4'
    baud_rate = 9600
    arduino = None

    try:
        arduino = serial.Serial(arduino_port, baud_rate)
        while True:
            rfid_data_full = arduino.readline().decode().strip()
            # Verificar si el valor es "Listo para leer etiquetas RFID..."
            if rfid_data_full != "Listo para leer etiquetas RFID...":
                # Extraer el UID eliminando el texto no deseado
                prefix = "Número de serie de la tarjeta RFID:"
                if rfid_data_full.startswith(prefix):
                    rfid_data = rfid_data_full[len(prefix):].strip().replace(" ", "")
                else:
                    rfid_data = rfid_data_full.replace(" ", "")

                if rfid_data:                    
                    print("RFID TAG: ", rfid_data)

                    # Verificar si el UID ya existe en la base de datos
                    existing_pechera = Pechera.objects.filter(id_pechera=rfid_data).first()

                    if existing_pechera:
                        print("UID ya existe en la base de datos")

                        # Generar la URL de la página 'lectura' con el ID de la pechera como parámetro
                        url = reverse('lectura') + f'?id={rfid_data}'

                        # Redirigir a la página 'lectura' con el ID de la pechera como parámetro en la URL
                        return redirect(url)
                    else:
                        # Notificar al usuario
                        print("UID no registrado en la base de datos")
                        return redirect('lectura_alerta')            

    except KeyboardInterrupt:
        print("Lectura de RFID detenida.")
    except Exception as e:
        print("Error: ", e)
    finally:
        if arduino is not None:
            arduino.close()

    return HttpResponse("Error en la lectura de datos, retroseda y recargue la pagina. IMPORTANTE: verifique que el dispositivo pueda leer información antes de comenzar la lectura.")

def editar(request):
    pechera = None
    
    if request.method == 'GET':
        id_pechera = request.GET.get("id")
        pechera = Pechera.objects.filter(id_pechera=id_pechera).first()

    # Obtener todas las plantas
    plantas = Planta.objects.all()

    parameters_choices = [
    ('1', 'Nuevo'),
    ('2', 'Perfectas condiciones'),
    ('3', 'Ligermante desgastanda'),
    ('4', 'Desgastada'),]

    opciones_indice = [('Si'),('No'),]
    context = {'pechera': pechera, 'plantas': plantas, 'parameters_choices': parameters_choices,'opciones_indice': opciones_indice}
    return render(request, 'app/editar.html', context)

def actualizar(request):
    pechera = Pechera.objects.filter(id_pechera=request.POST['id_pechera']).first()
    if pechera:
        pechera.talla = request.POST['talla']
        pechera.observaciones = request.POST['observaciones']
        pechera.planta = request.POST['planta']
        pechera.parameters = request.POST['parameters']
        pechera.indice_microbiologico = request.POST['indice_microbiologico']
        pechera.save()
    else:
        print("No se encontró la pechera con el UID especificado")
    return HttpResponseRedirect(f"/lectura?id={pechera.id_pechera}")

def exportar_a_excel(request):
    # Obtén los datos de la base de datos
    pecheras = Pechera.objects.all()
  
    # Crea un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active

    # Agrega encabezados a las columnas
    ws['A1'] = 'ID de Pechera'
    ws['B1'] = 'Planta'
    ws['C1'] = 'Fecha de Fabricación'
    ws['D1'] = 'Cantidad de Lavados'
    ws['E1'] = 'Talla'
    ws['F1'] = 'Observaciones'
    ws['G1'] = 'Parámetros'
    ws['H1'] = 'Índice Microbiológico'

    # Aplica estilos a los encabezados
    for cell in ws['1:1']:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

    # Aplica bordes a los encabezados
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.border = Border(bottom=Side(style='thin'))

    # Llena las filas con los datos de la base de datos
    row = 2
    for pechera in pecheras:
        ws.cell(row=row, column=1, value=pechera.id_pechera)
        ws.cell(row=row, column=2, value=pechera.get_planta_display())
        ws.cell(row=row, column=3, value=pechera.fecha_fabricacion.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=4, value=pechera.cantidad_lavados)
        ws.cell(row=row, column=5, value=pechera.talla)
        ws.cell(row=row, column=6, value=pechera.observaciones)
        ws.cell(row=row, column=7, value=pechera.get_parameters_display())
        ws.cell(row=row, column=8, value=pechera.indice_microbiologico)

        # Aplica estilos a las celdas
        for cell in ws[row]:
            cell.alignment = Alignment(horizontal='center')

        row += 1

    # Ajusta el ancho de las columnas para que se ajusten al contenido
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Crea una respuesta HTTP con el libro de Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="pecheras.xlsx"'
    wb.save(response)

    return response

def lavado(request):   
    pechera = None  # Inicialmente, no se muestra ninguna pechera
    if request.method == 'GET':
        id_pechera = request.GET.get("id")
        pechera = Pechera.objects.filter(id_pechera=id_pechera).first()
        if pechera:
            if pechera.eliminada:
                return redirect('lavado_alerta')  

    context = {'pechera': pechera}

    return render(request, 'app/lavado.html', context)

def lavado_alerta(request):   
    alerta = "UID no registrado en la base de datos" 
    pechera = None 
    context = {'pechera': pechera, 'alerta': alerta}
    return render(request, 'app/lavado.html', context)

def leer_lavado(request):    
    arduino_port = 'COM4'
    baud_rate = 9600
    arduino = None
    
    try:
        arduino = serial.Serial(arduino_port, baud_rate)
        
        while True:
            rfid_data_full = arduino.readline().decode().strip()
            # Verificar si el valor es "Listo para leer etiquetas RFID..."
            if rfid_data_full != "Listo para leer etiquetas RFID...":
                # Extraer el UID eliminando el texto no deseado
                prefix = "Número de serie de la tarjeta RFID:"
                if rfid_data_full.startswith(prefix):
                    rfid_data = rfid_data_full[len(prefix):].strip().replace(" ", "")
                else:
                    rfid_data = rfid_data_full.replace(" ", "")

                if rfid_data:                    
                    print("RFID TAG: ", rfid_data)

                    # Verificar si el UID ya existe en la base de datos
                    existing_pechera = Pechera.objects.filter(id_pechera=rfid_data).first()

                    if existing_pechera:
                        print("UID ya existe en la base de datos")

                        # Generar la URL de la página 'lectura' con el ID de la pechera como parámetro
                        url = reverse('lavado') + f'?id={rfid_data}'

                        # Redirigir a la página 'lectura' con el ID de la pechera como parámetro en la URL
                        return redirect(url)
                    else:
                        # Notificar al usuario
                        print("UID no registrado en la base de datos")
                        return redirect('lavado_alerta')            

    except KeyboardInterrupt:
        print("Lectura de RFID detenida.")
    except Exception as e:
        print("Error: ", e)
    finally:
        if arduino is not None:
            arduino.close()

    return HttpResponse("Error en la lectura de datos, retroseda y recargue la pagina. IMPORTANTE: verifique que el dispositivo pueda leer información antes de comenzar la lectura.")   

def registrar_lavado(request):
    pechera = None
    pechera = Pechera.objects.filter(id_pechera=request.POST['id_pechera']).first()
    if pechera:
        Lavado.objects.create(id_pechera = pechera.id_pechera, fecha_lavado = date.today())
        lavados = request.POST['lavados']
        pechera.cantidad_lavados = int(lavados)+1
        pechera.save()
    else:
        print("No se encontró la pechera con el UID especificado")
    return render(request, 'app/lavado.html')    

def eliminar(request):  
    pechera = None  # Inicialmente, no se muestra ninguna pechera
    if request.method == 'GET':
        id_pechera = request.GET.get("id")
        pechera = Pechera.objects.filter(id_pechera=id_pechera).first()
        if pechera:
            if pechera.eliminada:
                return redirect('eliminar_alerta')   

    context = {'pechera': pechera}

    return render(request, 'app/eliminar.html', context)

def eliminar_alerta(request):   
    alerta = "UID no registrado en la base de datos" 
    pechera = None 
    context = {'pechera': pechera, 'alerta': alerta}
    return render(request, 'app/eliminar.html', context)

def leer_eliminar(request):    
    arduino_port = 'COM4'
    baud_rate = 9600
    arduino = None
    
    try:
        arduino = serial.Serial(arduino_port, baud_rate)
        
        while True:
            rfid_data_full = arduino.readline().decode().strip()
            # Verificar si el valor es "Listo para leer etiquetas RFID..."
            if rfid_data_full != "Listo para leer etiquetas RFID...":
                # Extraer el UID eliminando el texto no deseado
                prefix = "Número de serie de la tarjeta RFID:"
                if rfid_data_full.startswith(prefix):
                    rfid_data = rfid_data_full[len(prefix):].strip().replace(" ", "")
                else:
                    rfid_data = rfid_data_full.replace(" ", "")

                if rfid_data:                    
                    print("RFID TAG: ", rfid_data)

                    # Verificar si el UID ya existe en la base de datos
                    existing_pechera = Pechera.objects.filter(id_pechera=rfid_data).first()

                    if existing_pechera:
                        print("UID ya existe en la base de datos")

                        # Generar la URL de la página 'lectura' con el ID de la pechera como parámetro
                        url = reverse('eliminar') + f'?id={rfid_data}'

                        # Redirigir a la página 'lectura' con el ID de la pechera como parámetro en la URL
                        return redirect(url)
                    else:
                        # Notificar al usuario
                        print("UID no registrado en la base de datos")
                        return redirect('eliminar_alerta')            

    except KeyboardInterrupt:
        print("Lectura de RFID detenida.")
    except Exception as e:
        print("Error: ", e)
    finally:
        if arduino is not None:
            arduino.close()

    return HttpResponse("Error en la lectura de datos, retroseda y recargue la pagina. IMPORTANTE: verifique que el dispositivo pueda leer información antes de comenzar la lectura.")   

def eliminar_pechera(request):
    pechera = None
    pechera = Pechera.objects.filter(id_pechera=request.POST['id_pechera']).first()
    if pechera:
        pechera.eliminada = True
        pechera.save()
    else:
        print("No se encontró la pechera con el UID especificado")
    return render(request, 'app/eliminar.html')  